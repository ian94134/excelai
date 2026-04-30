from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from build_web_ui_smoke_fixture import SALES_HEADERS, build_workbook
from tools_web_ui_smoke import (
    CASE_SETS,
    PLAIN_LANGUAGE_SMOKE_CASES,
    QUICK_ACTION_SMOKE_CASES,
    SMOKE_CASES,
)


def test_web_ui_smoke_fixture_has_expected_query_data(tmp_path):
    path = build_workbook(tmp_path / "web_ui_smoke_base.xlsx")

    wb = load_workbook(path, data_only=True)

    assert wb.sheetnames == ["Report", "SalesData"]
    assert wb.active.title == "Report"

    sales = wb["SalesData"]
    headers = [sales.cell(1, col).value for col in range(1, len(SALES_HEADERS) + 1)]
    assert headers == SALES_HEADERS

    north_amounts = [
        sales.cell(row, 5).value
        for row in range(2, sales.max_row + 1)
        if sales.cell(row, 2).value == "North"
    ]
    assert len(north_amounts) == 4
    assert sum(north_amounts) == 500


def test_web_ui_smoke_cases_are_short_and_ordered():
    names = [case.name for case in SMOKE_CASES]
    tools = [case.tool for case in SMOKE_CASES]

    assert names == [
        "fill_series_natural_phrase",
        "query_range_filter_sum",
        "beautify_range_business_blue",
        "write_range_undo_marker",
        "undo_last_blank_restore",
    ]
    assert tools == [
        "fill_series",
        "query_range",
        "beautify_range",
        "write_range",
        "undo_last",
    ]
    assert len({case.prompt for case in SMOKE_CASES}) == len(SMOKE_CASES)


def test_plain_language_smoke_cases_do_not_require_tool_names():
    names = [case.name for case in PLAIN_LANGUAGE_SMOKE_CASES]
    tools = [case.tool for case in PLAIN_LANGUAGE_SMOKE_CASES]

    assert names == [
        "plain_beautify_salesdata_report",
        "plain_query_north_amount_sum",
        "plain_write_report_aa1",
    ]
    assert tools == ["beautify_range", "query_range", "write_range"]
    assert set(CASE_SETS) == {"tool", "plain", "quick", "all"}
    assert CASE_SETS["plain"] == PLAIN_LANGUAGE_SMOKE_CASES
    assert CASE_SETS["quick"] == QUICK_ACTION_SMOKE_CASES
    assert CASE_SETS["all"] == SMOKE_CASES + PLAIN_LANGUAGE_SMOKE_CASES + QUICK_ACTION_SMOKE_CASES

    all_tool_names = {case.tool for case in SMOKE_CASES + PLAIN_LANGUAGE_SMOKE_CASES + QUICK_ACTION_SMOKE_CASES}
    for case in PLAIN_LANGUAGE_SMOKE_CASES + QUICK_ACTION_SMOKE_CASES:
        assert "呼叫" not in case.prompt
        assert "工具" not in case.prompt
        assert not any(tool_name in case.prompt for tool_name in all_tool_names)


def test_quick_action_smoke_cases_are_minimal():
    assert [case.name for case in QUICK_ACTION_SMOKE_CASES] == [
        "quick_action_beautify_report",
        "quick_action_summarize_data",
    ]
    assert [case.tool for case in QUICK_ACTION_SMOKE_CASES] == [
        "beautify_range",
        "summarize_range",
    ]
