"""Build the deterministic workbook used by the Web UI smoke test."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = PROJECT_ROOT / "tests" / "fixtures" / "web_ui_smoke_base.xlsx"

SALES_HEADERS = [
    "Date",
    "Region",
    "Product",
    "Quantity",
    "Amount",
    "Category",
    "Notes",
]

SALES_ROWS = [
    ["2026-01-01", "North", "Alpha", 12, 120, "A", "north apple"],
    ["2026-01-02", "South", "Beta", 24, 90, "B", "south beta"],
    ["2026-01-03", "East", "Gamma", 18, 75, "A", "east gamma"],
    ["2026-01-04", "North", "Delta", 8, 80, "C", "north delta"],
    ["2026-01-05", "West", "Epsilon", 6, 60, "B", "west epsilon"],
    ["2026-01-06", "North", "Zeta", 15, 150, "A", "north zeta"],
    ["2026-01-07", "South", "Eta", 11, 110, "C", "south eta"],
    ["2026-01-08", "East", "Theta", 5, 45, "B", "east theta"],
    ["2026-01-09", "North", "Iota", 14, 150, "A", "north iota"],
]


def build_workbook(output: Path = DEFAULT_OUTPUT) -> Path:
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    report = wb.active
    report.title = "Report"
    report["A1"] = "Smoke Area"
    report["A2"] = "Use K1:K5 for fill_series and Z1 for undo smoke."
    report["A1"].font = Font(bold=True, color="FFFFFF")
    report["A1"].fill = PatternFill("solid", fgColor="4472C4")
    report["A1"].alignment = Alignment(horizontal="center")
    report.column_dimensions["A"].width = 48

    sales = wb.create_sheet("SalesData")
    sales.append(SALES_HEADERS)
    for row in SALES_ROWS:
        sales.append(row)

    for col, width in {
        "A": 14,
        "B": 12,
        "C": 14,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 20,
    }.items():
        sales.column_dimensions[col].width = width

    wb.active = wb.sheetnames.index("Report")
    wb.save(output)
    return output


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Workbook path to create.",
    )
    args = parser.parse_args()

    path = build_workbook(args.output)
    print(f"created {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
